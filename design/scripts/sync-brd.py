#!/usr/bin/env python3
"""
BRD Validation + Aggregation Script

In v2.1 this script validates BRD ↔ md consistency.
In v2.2 it adds an aggregation mode: the BRD becomes a generated artifact
whose sheets are regenerated from md SSOT files.

Validation sources:
  - story-map.md (DS-NNN IDs)
  - business-rules-register.md (BR-NN IDs)
  - screen-inventory.md (story-to-screen mapping)
  - release-slices.md (story IDs)
  - BRD.xlsx "Manifest" sheet (artifact version freshness; migrated from BRD_manifest.md in v2.1)

Aggregation sources (v2.2, --regenerate mode):
  - design/05_STORIES/story-map.md             → User Stories sheet (cols A–H)
  - design/04_PROCESS_FLOWS/business-rules-register.md → inline-expanded into AC cells via [BR-NN] tags
  - design/06_INFORMATION_ARCHITECTURE/rbac.md          → RBAC sheet
  - design/06_INFORMATION_ARCHITECTURE/notifications.md → Notification Mapping sheet
  - design/06_INFORMATION_ARCHITECTURE/data-dictionary.md → Data Fields sheet
  - design/09_CONTENT/terminology.md (LOV section)        → LOV sheet
  - design/06_INFORMATION_ARCHITECTURE/screen-inventory.md → Feature/Touchpoint col C (reverse lookup)

Aggregation rules:
  - Story-origin AC bullets are untagged (implied origin); downstream modes
    append tagged bullets [BR-NN], [STATE], [BEHAVIOR], [A11Y], [CANVAS], [NOTIF-NNN], [FLOW].
  - [BR-NN] tags are *inline-expanded* in the AC cell — the BR text from
    business-rules-register.md is written into the cell so the BRD reader
    sees the combined AC + BR view per story.
  - Other tag families remain as in-cell references (no expansion).
  - Priority/release columns are out of regeneration scope. Existing values
    in those columns are preserved; the script warns if they look stale.
  - If navigation-model.md still contains a role-feature matrix and rbac.md
    also exists, rbac.md wins and the script warns about the duplicate.

Usage:
    python design/scripts/sync-brd.py [project_root]                   (validate)
    python design/scripts/sync-brd.py --regenerate [project_root]      (regenerate sheets from md SSOTs)
    python design/scripts/sync-brd.py --dry-run [project_root]         (validate aggregation without writing)
    python design/scripts/sync-brd.py --migrate-manifest [project_root]

Exit code 1 if errors found, 0 otherwise.
"""

import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def resolve_root(argv):
    """Resolve project root from CLI arg or default to ../../ relative to script."""
    if len(argv) > 1:
        return os.path.abspath(argv[1])
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(script_dir, '..', '..'))


def read_file(path):
    """Read a file and return its contents, or None if missing."""
    if not os.path.isfile(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


def extract_ids(text, pattern):
    """Extract all unique IDs matching a regex pattern from text."""
    if text is None:
        return set()
    return set(re.findall(pattern, text))


def read_brd(brd_path):
    """
    Read the BRD User Stories sheet.
    Returns (story_ids, ac_text_by_row, feature_by_row) where:
      - story_ids: set of DS-NNN from column D
      - ac_text_by_row: dict {row: acceptance_criteria_text} from column F
      - feature_by_row: dict {row: feature_value} from column C
    Data rows start at row 4 (row 3 is headers).
    """
    if not os.path.isfile(brd_path):
        return None, None, None

    wb = load_workbook(brd_path, data_only=True)

    # Find the User Stories sheet (case-insensitive partial match)
    sheet = None
    for name in wb.sheetnames:
        if 'user stor' in name.lower():
            sheet = wb[name]
            break

    if sheet is None:
        return None, None, None

    story_ids = set()
    ac_text_by_row = {}
    feature_by_row = {}

    for row_num in range(4, sheet.max_row + 1):
        # Column D = story ID
        cell_d = sheet.cell(row=row_num, column=4).value
        if cell_d:
            ids = re.findall(r'DS-\d+', str(cell_d))
            story_ids.update(ids)

        # Column F = acceptance criteria
        cell_f = sheet.cell(row=row_num, column=6).value
        if cell_f:
            ac_text_by_row[row_num] = str(cell_f)

        # Column C = feature/touchpoint
        cell_c = sheet.cell(row=row_num, column=3).value
        feature_by_row[row_num] = cell_c

    wb.close()
    return story_ids, ac_text_by_row, feature_by_row


def check_story_map_to_brd(story_map_ids, brd_ids):
    """Check 1: Every DS-NNN in story-map.md exists in BRD col D."""
    missing = sorted(story_map_ids - brd_ids, key=lambda x: int(x.split('-')[1]))
    return missing


def check_brd_to_story_map(brd_ids, story_map_ids):
    """Check 2: Every DS-NNN in BRD col D exists in story-map.md."""
    missing = sorted(brd_ids - story_map_ids, key=lambda x: int(x.split('-')[1]))
    return missing


def check_br_ids_in_brd(ac_text_by_row, br_ids):
    """Check 3: Every BR-NN in BRD col F exists in business-rules-register.md."""
    brd_br_ids = set()
    for text in ac_text_by_row.values():
        brd_br_ids.update(re.findall(r'BR-\d+', text))

    missing = sorted(brd_br_ids - br_ids, key=lambda x: int(x.split('-')[1]))
    return missing, brd_br_ids


def check_feature_coverage(brd_ids, feature_by_row, brd_path):
    """Check 4: Warn if a story has no Feature/Touchpoint value in col C."""
    if not os.path.isfile(brd_path):
        return []

    wb = load_workbook(brd_path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if 'user stor' in name.lower():
            sheet = wb[name]
            break

    if sheet is None:
        wb.close()
        return []

    warnings = []
    for row_num in range(4, sheet.max_row + 1):
        cell_d = sheet.cell(row=row_num, column=4).value
        cell_c = sheet.cell(row=row_num, column=3).value
        if cell_d and re.search(r'DS-\d+', str(cell_d)):
            if not cell_c or str(cell_c).strip() == '':
                ids = re.findall(r'DS-\d+', str(cell_d))
                for sid in ids:
                    warnings.append(f"{sid} (row {row_num}) has no Feature/Touchpoint value")

    wb.close()
    return warnings


def check_manifest_freshness_md(manifest_path):
    """Legacy (pre-v2.1): read BRD_manifest.md and warn if any mode shows '—'."""
    content = read_file(manifest_path)
    if content is None:
        return None, []

    warnings = []
    for line in content.splitlines():
        line = line.strip()
        if line.startswith('|') and not line.startswith('|---') and not line.startswith('| Mode'):
            cols = [c.strip() for c in line.split('|')]
            if len(cols) >= 5:
                mode = cols[1]
                last_contributed = cols[2]
                if last_contributed == '—' or last_contributed == '':
                    warnings.append(f"Mode '{mode}' has never contributed to BRD")

    return content, warnings


def check_manifest_freshness_sheet(brd_path):
    """v2.1+: read BRD.xlsx "Manifest" sheet and warn if any mode shows '—'."""
    if not os.path.isfile(brd_path):
        return None, []

    wb = load_workbook(brd_path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.lower() == 'manifest':
            sheet = wb[name]
            break

    if sheet is None:
        wb.close()
        return None, []

    warnings = []
    # Expected columns: Mode | Last contributed | Artifact version | Stories touched | Notes
    # Row 1 = header
    for row_num in range(2, sheet.max_row + 1):
        mode = sheet.cell(row=row_num, column=1).value
        last_contributed = sheet.cell(row=row_num, column=2).value
        if not mode:
            continue
        if last_contributed is None or str(last_contributed).strip() in ('—', ''):
            warnings.append(f"Mode '{mode}' has never contributed to BRD")

    wb.close()
    return True, warnings


def migrate_manifest_to_sheet(brd_path, manifest_path):
    """
    One-shot v2.1 migration: move design/BRD_manifest.md contents into a
    "Manifest" sheet inside BRD.xlsx, then delete the .md file.
    """
    if not os.path.isfile(brd_path):
        print(f"✗ BRD not found at {brd_path}. Nothing to migrate into.")
        return 1
    if not os.path.isfile(manifest_path):
        print(f"ℹ BRD_manifest.md not found at {manifest_path} — already migrated or never existed.")
        return 0

    content = read_file(manifest_path)
    if content is None:
        print(f"✗ Could not read {manifest_path}")
        return 1

    # Parse markdown table rows
    rows = []
    for line in content.splitlines():
        line = line.strip()
        if line.startswith('|') and not line.startswith('|---') and not line.startswith('| Mode'):
            cols = [c.strip() for c in line.split('|')]
            # Leading and trailing '' from split — drop them
            cols = [c for c in cols if c != '']
            if cols:
                rows.append(cols)

    if not rows:
        print(f"⚠ No table rows parsed from {manifest_path}. Migration aborted — inspect manually.")
        return 1

    # Extract header from the markdown too
    header = None
    for line in content.splitlines():
        line = line.strip()
        if line.startswith('| Mode'):
            header = [c.strip() for c in line.split('|') if c.strip() != '']
            break
    if header is None:
        header = ['Mode', 'Last contributed', 'Artifact version', 'Stories touched', 'Notes']

    wb = load_workbook(brd_path)
    if 'Manifest' in wb.sheetnames:
        print(f"✗ BRD already has a 'Manifest' sheet. Aborting to avoid overwrite.")
        print(f"  Resolve manually: review both sources and delete one.")
        wb.close()
        return 1

    sheet = wb.create_sheet('Manifest')
    sheet.append(header)
    for row in rows:
        # Pad/truncate to header length
        padded = row + [''] * (len(header) - len(row))
        sheet.append(padded[:len(header)])

    wb.save(brd_path)
    wb.close()

    os.remove(manifest_path)
    print(f"✓ Migrated {len(rows)} row(s) from BRD_manifest.md into BRD.xlsx → Manifest sheet.")
    print(f"✓ Deleted {manifest_path}")
    print(f"\nReview with `git status` and commit when ready.")
    return 0


def check_release_slices_to_brd(release_ids, brd_ids):
    """Check 6: Every DS-NNN in release-slices.md exists in BRD col D."""
    missing = sorted(release_ids - brd_ids, key=lambda x: int(x.split('-')[1]))
    return missing


# ============================================================================
# v2.2 — Aggregation: regenerate BRD sheets from md SSOT files
# ============================================================================

# Tag families recognised in AC bullets. Only [BR-NN] is inline-expanded;
# others remain as in-cell references for the BRD reader.
TAG_PATTERNS = {
    'BR': re.compile(r'\[BR-(\d+)\]'),
    'STATE': re.compile(r'\[STATE\]'),
    'BEHAVIOR': re.compile(r'\[BEHAVIOR\]'),
    'A11Y': re.compile(r'\[A11Y\]'),
    'CANVAS': re.compile(r'\[CANVAS\]'),
    'FLOW': re.compile(r'\[FLOW\]'),
    'NOTIF': re.compile(r'\[NOTIF-(\d+)\]'),
}


def parse_business_rules(br_text):
    """
    Parse business-rules-register.md into {BR-NN: rule_text}.
    Recognises markdown table rows of the form:
      | BR-01 | <description> | ... |
    or H3 sections of the form:
      ### BR-01 — <title>
      <body>
    Returns a dict mapping each id to its rule text (single line, table form
    preferred when both exist).
    """
    if not br_text:
        return {}

    rules = {}

    # Table-row form: | BR-NN | description | ...
    table_pattern = re.compile(r'^\|\s*(BR-\d+)\s*\|\s*([^|]+?)\s*\|', re.MULTILINE)
    for match in table_pattern.finditer(br_text):
        rules[match.group(1)] = match.group(2).strip()

    # H3 section form: ### BR-NN — title
    section_pattern = re.compile(
        r'^###\s*(BR-\d+)\s*[—–-]\s*([^\n]+?)\s*$\n+(.*?)(?=^###|\Z)',
        re.MULTILINE | re.DOTALL,
    )
    for match in section_pattern.finditer(br_text):
        bid, title, body = match.group(1), match.group(2).strip(), match.group(3).strip()
        if bid not in rules:
            # Use first non-blank body line as the rule text; fall back to title.
            first_line = next((line.strip() for line in body.splitlines() if line.strip()), title)
            rules[bid] = first_line

    return rules


def parse_story_map(story_map_text):
    """
    Parse story-map.md into a list of story dicts:
      [{'id': 'DS-001', 'title': '...', 'persona': '...', 'narrative': '...', 'ac': [...]}]
    Each AC entry is the raw bullet text (preserving tags). Recognises stories
    introduced by an H3 of the form '### DS-NNN — Story title' followed by an
    AC list under '**Acceptance criteria**' or 'AC:' or a fenced section.
    """
    if not story_map_text:
        return []

    stories = []
    # Split on H3 story headers
    story_pattern = re.compile(
        r'^###\s*(DS-\d+)\s*[—–-]\s*([^\n]+?)\s*$',
        re.MULTILINE,
    )
    matches = list(story_pattern.finditer(story_map_text))

    for i, match in enumerate(matches):
        sid = match.group(1)
        title = match.group(2).strip()
        body_start = match.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(story_map_text)
        body = story_map_text[body_start:body_end]

        # Extract persona (first 'Persona:' or '**Persona**' line)
        persona_match = re.search(r'(?:^\*\*Persona\*\*:?|^Persona:)\s*(.+)$', body, re.MULTILINE)
        persona = persona_match.group(1).strip() if persona_match else ''

        # Extract narrative (As a X, I want Y, so that Z)
        narr_match = re.search(r'^[*_]*As an?\s.+', body, re.MULTILINE)
        narrative = narr_match.group(0).strip() if narr_match else ''

        # Extract AC bullets — block under 'Acceptance criteria' header
        ac = []
        ac_block_match = re.search(
            r'(?:\*\*Acceptance criteria\*\*|^####?\s*Acceptance criteria|^AC:)(.*?)(?=^####?\s|\Z)',
            body,
            re.MULTILINE | re.DOTALL,
        )
        if ac_block_match:
            for line in ac_block_match.group(1).splitlines():
                line = line.strip()
                if line.startswith('-') or line.startswith('*'):
                    ac.append(line.lstrip('-* ').strip())

        stories.append({
            'id': sid,
            'title': title,
            'persona': persona,
            'narrative': narrative,
            'ac': ac,
        })

    return stories


def expand_ac_bullets(ac_bullets, br_rules):
    """
    Inline-expand [BR-NN] tags in each AC bullet by appending the BR text in
    parentheses. Other tag families remain as bare references.

    Returns the expanded text as a single string with one bullet per line
    (prefixed with '• '), suitable for writing into a single Excel cell.
    """
    expanded = []
    for bullet in ac_bullets:
        text = bullet
        for match in TAG_PATTERNS['BR'].finditer(bullet):
            bid = f"BR-{match.group(1)}"
            rule = br_rules.get(bid)
            if rule:
                # Replace the bare tag with tag + inline text
                text = text.replace(match.group(0), f"{match.group(0)} ({rule})")
        expanded.append(f"• {text}")
    return '\n'.join(expanded)


def parse_md_table(text, header_keywords):
    """
    Generic markdown-table parser. Finds the first table whose header row
    contains all of `header_keywords` (case-insensitive substring match) and
    returns a list of row dicts keyed by lowercased header.
    """
    if not text:
        return []

    rows = []
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith('|') and i + 1 < len(lines):
            header_cells = [c.strip().lower() for c in line.strip('|').split('|')]
            if all(any(kw in h for h in header_cells) for kw in header_keywords):
                # Skip separator row
                i += 2
                while i < len(lines) and lines[i].strip().startswith('|'):
                    cells = [c.strip() for c in lines[i].strip('|').split('|')]
                    if len(cells) == len(header_cells):
                        rows.append(dict(zip(header_cells, cells)))
                    i += 1
                return rows
        i += 1
    return rows


def regenerate_user_stories_sheet(wb, stories, br_rules, screen_inventory_text):
    """
    Regenerate the User Stories sheet (cols A–H) from story-map.md.
    Preserves cols I+ (priority/release/notes — out of scope per v2.2).
    Col C (Feature/Touchpoint) is reverse-looked-up from screen-inventory.
    """
    sheet = None
    for name in wb.sheetnames:
        if 'user stor' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        return 0

    # Build reverse lookup: DS-NNN → list of screen names from screen-inventory
    story_to_screens = {}
    if screen_inventory_text:
        # Look for "Stories served: DS-001, DS-005" patterns under each screen
        screen_pattern = re.compile(
            r'^###?\s*([^\n]+?)\s*$\n.*?Stories served:?\s*([^\n]+)',
            re.MULTILINE | re.DOTALL,
        )
        for match in screen_pattern.finditer(screen_inventory_text):
            screen_name = match.group(1).strip()
            for sid in re.findall(r'DS-\d+', match.group(2)):
                story_to_screens.setdefault(sid, []).append(screen_name)

    written = 0
    # Data rows start at row 4 (row 3 is headers)
    # Map existing story IDs to row numbers so we update in place
    existing_rows = {}
    for row_num in range(4, sheet.max_row + 1):
        cell_d = sheet.cell(row=row_num, column=4).value
        if cell_d:
            for sid in re.findall(r'DS-\d+', str(cell_d)):
                existing_rows[sid] = row_num

    next_new_row = sheet.max_row + 1

    for story in stories:
        row = existing_rows.get(story['id'], next_new_row)
        if story['id'] not in existing_rows:
            next_new_row += 1

        # Col C = Feature/Touchpoint (reverse lookup)
        screens = story_to_screens.get(story['id'], [])
        if screens:
            sheet.cell(row=row, column=3).value = ', '.join(screens)
        # Col D = Story ID
        sheet.cell(row=row, column=4).value = story['id']
        # Col E = Title / narrative
        sheet.cell(row=row, column=5).value = story['narrative'] or story['title']
        # Col F = Acceptance Criteria (with [BR-NN] inline-expanded)
        sheet.cell(row=row, column=6).value = expand_ac_bullets(story['ac'], br_rules)
        # Col G = Persona
        sheet.cell(row=row, column=7).value = story['persona']
        # Col H = Title (raw)
        sheet.cell(row=row, column=8).value = story['title']
        # Cols I+ left untouched (priority/release/notes — out of scope)

        written += 1

    return written


def regenerate_simple_sheet(wb, sheet_name, header_keywords, md_text, column_order):
    """
    Generic generator: replace `sheet_name` with rows parsed from a md table
    matching `header_keywords`. `column_order` is the list of header keys (in
    order) to write to the sheet's columns A, B, C, …
    """
    rows = parse_md_table(md_text, header_keywords)
    if not rows:
        return 0

    # Drop existing sheet if present, recreate
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    sheet = wb.create_sheet(sheet_name)

    # Header row from column_order (Title Case)
    sheet.append([key.replace('_', ' ').title() for key in column_order])

    written = 0
    for row in rows:
        sheet.append([row.get(col, '') for col in column_order])
        written += 1
    return written


def regenerate(brd_path, sources, dry_run=False):
    """
    Top-level aggregation entry point. Reads md SSOT sources, regenerates each
    BRD sheet, and writes the workbook. Returns (errors, warnings).
    """
    errors = 0
    warnings = 0

    if not os.path.isfile(brd_path):
        print(f"✗ BRD not found at {brd_path}. Run from a project root with design/BRD.xlsx present.")
        return 1, 0

    print(f"=== BRD Regeneration (v2.2 aggregation) ===")
    print(f"Mode: {'DRY RUN' if dry_run else 'WRITE'}")
    print(f"BRD:  {brd_path}\n")

    # ---- Load sources ----
    story_map_text = read_file(sources['story_map'])
    br_text = read_file(sources['business_rules'])
    rbac_text = read_file(sources['rbac'])
    nav_model_text = read_file(sources['navigation_model'])
    notif_text = read_file(sources['notifications'])
    data_dict_text = read_file(sources['data_dictionary'])
    terminology_text = read_file(sources['terminology'])
    screen_inv_text = read_file(sources['screen_inventory'])

    # ---- Parse upstream ----
    br_rules = parse_business_rules(br_text)
    stories = parse_story_map(story_map_text)
    print(f"Parsed {len(stories)} stories, {len(br_rules)} business rules.\n")

    # ---- Conflict check: navigation-model.md still has role-feature matrix? ----
    if rbac_text and nav_model_text:
        # Heuristic — if navigation-model.md has a table whose header includes 'Role' AND a feature/screen-ish keyword
        if parse_md_table(nav_model_text, ['role', 'feature']) or \
           parse_md_table(nav_model_text, ['role', 'screen']):
            print("⚠ rbac.md exists AND navigation-model.md still contains a role-feature matrix.")
            print("  rbac.md wins; remove the duplicate from navigation-model.md.\n")
            warnings += 1

    if dry_run:
        wb = load_workbook(brd_path, data_only=False)
    else:
        wb = load_workbook(brd_path)

    # ---- User Stories sheet ----
    n = regenerate_user_stories_sheet(wb, stories, br_rules, screen_inv_text)
    print(f"  User Stories: {n} row(s) written (cols A–H; AC inline-expanded with BR text).")

    # ---- RBAC sheet ----
    if rbac_text:
        n = regenerate_simple_sheet(wb, 'RBAC', ['role'], rbac_text,
                                     column_order=['screen', 'role', 'access', 'notes'])
        print(f"  RBAC: {n} row(s) written from rbac.md.")
    else:
        print("  RBAC: skipped — design/06_INFORMATION_ARCHITECTURE/rbac.md not found.")
        warnings += 1

    # ---- Notification Mapping sheet ----
    if notif_text:
        n = regenerate_simple_sheet(wb, 'Notification Mapping', ['notif'], notif_text,
                                     column_order=['notif id', 'channel', 'recipient', 'copy template', 'trigger event'])
        print(f"  Notification Mapping: {n} row(s) written from notifications.md.")
    else:
        print("  Notification Mapping: skipped — design/06_INFORMATION_ARCHITECTURE/notifications.md not found.")
        warnings += 1

    # ---- Data Fields sheet ----
    if data_dict_text:
        n = regenerate_simple_sheet(wb, 'Data Fields', ['field'], data_dict_text,
                                     column_order=['field id', 'type', 'validation', 'format', 'source screen'])
        print(f"  Data Fields: {n} row(s) written from data-dictionary.md.")
    else:
        print("  Data Fields: skipped — design/06_INFORMATION_ARCHITECTURE/data-dictionary.md not found.")
        warnings += 1

    # ---- LOV sheet ----
    if terminology_text:
        # Look for an LOV section then parse its tables
        lov_match = re.search(r'^##\s*LOV.*?(?=^##\s|\Z)', terminology_text, re.MULTILINE | re.DOTALL)
        if lov_match:
            n = regenerate_simple_sheet(wb, 'LOV', ['value'], lov_match.group(0),
                                         column_order=['group', 'value', 'label', 'notes'])
            print(f"  LOV: {n} row(s) written from terminology.md (LOV section).")
        else:
            print("  LOV: skipped — terminology.md has no '## LOV' section.")
            warnings += 1
    else:
        print("  LOV: skipped — design/09_CONTENT/terminology.md not found.")
        warnings += 1

    # ---- Save ----
    if not dry_run:
        wb.save(brd_path)
        print(f"\n✓ BRD regenerated → {brd_path}")
        print(f"  (priority/release columns preserved; manual edits to those cols are not overwritten)")
    else:
        print(f"\n[DRY RUN] No changes written.")

    wb.close()

    print(f"\n=== Summary: {errors} errors, {warnings} warnings ===")
    return errors, warnings


def main():
    flags = {'--migrate-manifest', '--regenerate', '--dry-run'}
    argv = [a for a in sys.argv if a not in flags]
    migrate_flag = '--migrate-manifest' in sys.argv
    regenerate_flag = '--regenerate' in sys.argv
    dry_run_flag = '--dry-run' in sys.argv

    root = resolve_root(argv)

    brd_path = os.path.join(root, 'design', 'BRD.xlsx')
    story_map_path = os.path.join(root, 'design', '05_STORIES', 'story-map.md')
    br_register_path = os.path.join(root, 'design', '04_PROCESS_FLOWS', 'business-rules-register.md')
    screen_inv_path = os.path.join(root, 'design', '06_INFORMATION_ARCHITECTURE', 'screen-inventory.md')
    nav_model_path = os.path.join(root, 'design', '06_INFORMATION_ARCHITECTURE', 'navigation-model.md')
    rbac_path = os.path.join(root, 'design', '06_INFORMATION_ARCHITECTURE', 'rbac.md')
    notif_path = os.path.join(root, 'design', '06_INFORMATION_ARCHITECTURE', 'notifications.md')
    data_dict_path = os.path.join(root, 'design', '06_INFORMATION_ARCHITECTURE', 'data-dictionary.md')
    terminology_path = os.path.join(root, 'design', '09_CONTENT', 'terminology.md')
    release_slices_path = os.path.join(root, 'design', '05_STORIES', 'release-slices.md')
    manifest_path = os.path.join(root, 'design', 'BRD_manifest.md')

    if migrate_flag:
        print("=== BRD Manifest Migration (v2.1) ===\n")
        rc = migrate_manifest_to_sheet(brd_path, manifest_path)
        sys.exit(rc)

    if regenerate_flag or dry_run_flag:
        sources = {
            'story_map': story_map_path,
            'business_rules': br_register_path,
            'rbac': rbac_path,
            'navigation_model': nav_model_path,
            'notifications': notif_path,
            'data_dictionary': data_dict_path,
            'terminology': terminology_path,
            'screen_inventory': screen_inv_path,
        }
        errors, _warnings = regenerate(brd_path, sources, dry_run=dry_run_flag)
        sys.exit(1 if errors > 0 else 0)

    errors = 0
    warnings = 0

    print("=== BRD Validation Report ===\n")

    # ---- Load BRD ----
    brd_ids, ac_text_by_row, feature_by_row = read_brd(brd_path)
    if brd_ids is None:
        print(f"WARNING: BRD not found at {brd_path} — skipping all BRD checks\n")
        print("=== Summary: 0 errors, 1 warning ===")
        sys.exit(0)

    # ---- Load story map ----
    story_map_text = read_file(story_map_path)
    story_map_ids = extract_ids(story_map_text, r'DS-\d+')

    # ---- Load business rules register ----
    br_text = read_file(br_register_path)
    br_ids = extract_ids(br_text, r'BR-\d+')

    # ---- Load release slices ----
    release_text = read_file(release_slices_path)
    release_ids = extract_ids(release_text, r'DS-\d+')

    # ---- Check 1: Story Map → BRD ----
    print("--- 1. Story Map → BRD ---")
    if story_map_text is None:
        print(f"WARNING: story-map.md not found at {story_map_path} — skipping\n")
        warnings += 1
    else:
        missing = check_story_map_to_brd(story_map_ids, brd_ids)
        if not missing:
            print(f"✓ All {len(story_map_ids)} stories in story-map.md exist in BRD\n")
        else:
            for sid in missing:
                print(f"✗ {sid} not found in BRD")
                errors += 1
            print()

    # ---- Check 2: BRD → Story Map ----
    print("--- 2. BRD → Story Map ---")
    if story_map_text is None:
        print(f"WARNING: story-map.md not found — skipping\n")
        warnings += 1
    else:
        missing = check_brd_to_story_map(brd_ids, story_map_ids)
        if not missing:
            print(f"✓ All {len(brd_ids)} stories in BRD exist in story-map.md\n")
        else:
            for sid in missing:
                print(f"✗ {sid} in BRD not found in story-map.md")
                errors += 1
            print()

    # ---- Check 3: BR-NN in BRD AC → business-rules-register ----
    print("--- 3. BRD Acceptance Criteria BR Tags → Business Rules Register ---")
    if br_text is None:
        print(f"WARNING: business-rules-register.md not found at {br_register_path} — skipping\n")
        warnings += 1
    elif ac_text_by_row is None:
        print("WARNING: No acceptance criteria data in BRD — skipping\n")
        warnings += 1
    else:
        missing, brd_br_ids = check_br_ids_in_brd(ac_text_by_row, br_ids)
        if not brd_br_ids:
            print("✓ No BR-NN tags found in BRD acceptance criteria (none to validate)\n")
        elif not missing:
            print(f"✓ All {len(brd_br_ids)} BR tags in BRD exist in business-rules-register.md\n")
        else:
            for bid in missing:
                print(f"✗ {bid} in BRD acceptance criteria not found in business-rules-register.md")
                errors += 1
            print()

    # ---- Check 4: Feature/Touchpoint coverage ----
    print("--- 4. Feature/Touchpoint Coverage ---")
    feat_warnings = check_feature_coverage(brd_ids, feature_by_row, brd_path)
    if not feat_warnings:
        print(f"✓ All stories in BRD have a Feature/Touchpoint value\n")
    else:
        for w in feat_warnings:
            print(f"⚠ {w}")
            warnings += 1
        print()

    # ---- Check 5: Manifest freshness ----
    # v2.1+: prefer the Manifest sheet inside BRD.xlsx; fall back to BRD_manifest.md
    print("--- 5. Manifest Freshness ---")
    sheet_present, sheet_warnings = check_manifest_freshness_sheet(brd_path)
    if sheet_present:
        if not sheet_warnings:
            print("✓ All modes have contributed to BRD (Manifest sheet)\n")
        else:
            for w in sheet_warnings:
                print(f"⚠ {w}")
                warnings += 1
            print()
    else:
        manifest_content, manifest_warnings = check_manifest_freshness_md(manifest_path)
        if manifest_content is None:
            print(f"WARNING: No Manifest sheet in BRD.xlsx and no BRD_manifest.md found — skipping")
            print(f"         (v2.1 migrates the manifest into the xlsx; run sync-brd.py --migrate-manifest if upgrading)\n")
            warnings += 1
        else:
            print(f"ℹ Reading legacy BRD_manifest.md. Run --migrate-manifest to consolidate into BRD.xlsx.")
            if not manifest_warnings:
                print("✓ All modes have contributed to BRD\n")
            else:
                for w in manifest_warnings:
                    print(f"⚠ {w}")
                    warnings += 1
                print()

    # ---- Check 6: Release Slices → BRD ----
    print("--- 6. Release Slices → BRD ---")
    if release_text is None:
        print(f"WARNING: release-slices.md not found at {release_slices_path} — skipping\n")
        warnings += 1
    else:
        missing = check_release_slices_to_brd(release_ids, brd_ids)
        if not release_ids:
            print("✓ No DS-NNN IDs found in release-slices.md (none to validate)\n")
        elif not missing:
            print(f"✓ All {len(release_ids)} stories in release-slices.md exist in BRD\n")
        else:
            for sid in missing:
                print(f"✗ {sid} in release-slices.md not found in BRD")
                errors += 1
            print()

    # ---- Summary ----
    print(f"=== Summary: {errors} errors, {warnings} warnings ===")
    sys.exit(1 if errors > 0 else 0)


if __name__ == '__main__':
    main()
