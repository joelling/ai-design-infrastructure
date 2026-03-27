"""
populate_brd.py — Populate design/BRD.xlsx with Keppel9D RFA Comment Response workflow data.

Sources consumed:
  design/05_STORIES/rfa-response-story-map.md       (story-map@v1)
  design/04_PROCESS_FLOWS/business-rules-register.md (business-rules-register@v1)
  design/06_INFORMATION_ARCHITECTURE/screen-inventory.md   (screen-inventory@v2)
  design/06_INFORMATION_ARCHITECTURE/navigation-model.md   (navigation-model@v2)
  design/06_INFORMATION_ARCHITECTURE/entity-taxonomy.md    (entity-taxonomy@v1)
  design/07_INTERACTION/index.md                           (index@v1)

Sheets populated:
  User Stories      — 26 rows, cols A–H only (cols J/L/N/P have pre-loaded VLOOKUP formulas)
  RBAC              — 10 screen rows × 5 role columns (role names replace placeholder headers)
  Data Fields       — 49 field rows across 10 entities
  LOV               — 33 rows across 8 enum groups
  Notification Mapping — 6 notification rows
"""

import os
import subprocess
import sys
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BRD_PATH = os.path.join(BASE, "BRD.xlsx")
RECALC = os.path.join(BASE, "scripts", "recalc.py")

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------

STORIES = [
    {
        "sno": 1,
        "epic": "Receive notification",
        "feature": "Receive alert",
        "screen": "P-03 Pending Reviews",
        "id": "DS-001",
        "story": (
            "When a new RFA revision batch has been processed and AI drafts are ready, "
            "I want to be notified without having to check manually, "
            "so I can start reviewing without delay."
        ),
        "ac": (
            "[STORY] Notification is received without initiating it; "
            "it contains enough context to identify the revision and RFA"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 2,
        "epic": "Receive notification",
        "feature": "Receive alert",
        "screen": "P-03 Pending Reviews",
        "id": "DS-002",
        "story": (
            "When I receive a batch notification, I want to know the scale of the review task "
            "at a glance (number of comments, number flagged), "
            "so I can plan my review time before opening the batch."
        ),
        "ac": (
            "[STORY] Count of total comments and flagged comments is visible in or with the notification"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 3,
        "epic": "Receive notification",
        "feature": "Know revision/RFA before opening",
        "screen": "P-03 Pending Reviews",
        "id": "DS-003",
        "story": (
            "When I am notified of a new batch, I want to confirm which RFA and revision "
            "the batch belongs to before I open it, so I can orient to the project context."
        ),
        "ac": (
            "[STORY] RFA reference, revision number, and submission date are identifiable "
            "from the notification"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 4,
        "epic": "Orient to batch",
        "feature": "See comment list with indicators",
        "screen": "P-02 Project Space",
        "id": "DS-004",
        "story": (
            "When I open a batch, I want to see all SO comments in the batch with a status "
            "indicator for each (AI draft available, cross-reference, flagged), "
            "so I can scan the full scope before reviewing individual drafts."
        ),
        "ac": (
            "[STORY] All comments are visible in one place; "
            "status is indicated without opening each comment"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 5,
        "epic": "Orient to batch",
        "feature": "Identify cross-references",
        "screen": "P-02 Project Space",
        "id": "DS-005",
        "story": (
            "When I open a batch, I want to immediately identify which comments have been "
            "short-circuited as cross-references (no AI generation), "
            "so I can treat them differently from full AI drafts."
        ),
        "ac": (
            "[STORY] Cross-reference comments are visually distinguishable; "
            "the referenced comment number is shown\n"
            "[BR-03] If comment intent matches an existing approved response above similarity "
            "threshold, system routes as cross-reference without AI generation; "
            "referenced comment number is shown"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 6,
        "epic": "Orient to batch",
        "feature": "Identify high-priority flags",
        "screen": "P-02 Project Space",
        "id": "DS-006",
        "story": (
            "When I open a batch, I want to see which comments are carrying high-priority flags "
            "(compliance bias warning, max-iterations alert, low confidence), "
            "so I can review the most commercially consequential drafts first."
        ),
        "ac": (
            "[STORY] Flagged comments are identifiable from the batch overview; "
            "compliance bias flags are the highest visual priority"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 7,
        "epic": "Orient to batch",
        "feature": "Identify high-priority flags",
        "screen": "P-02 Project Space",
        "id": "DS-007",
        "story": (
            "When I see a compliance-bias flag on a comment, I want to understand what kind "
            "of compliance issue was detected before I open the draft, "
            "so I can approach the review with the right scrutiny."
        ),
        "ac": (
            "[STORY] Compliance flag is labelled in a way that communicates the risk "
            "(AI agreed to change request rather than defending design), not just that a flag exists\n"
            "[BR-06] Compliance flag communicates that AI agreed to a change request rather than "
            "defending design; risk consequence is visible before opening draft\n"
            "[ASSUMPTION] BR-06 compliance bias detection heuristics — exact trigger conditions TBC"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 8,
        "epic": "Orient to batch",
        "feature": "Batch confidence profile",
        "screen": "P-02 Project Space",
        "id": "DS-008",
        "story": (
            "When I open a batch, I want to see an overall confidence profile for the batch "
            "(proportion of high/medium/low confidence drafts), "
            "so I can calibrate the effort I need before I begin."
        ),
        "ac": (
            "[STORY] Confidence distribution is visible at batch level; "
            "does not require opening individual drafts\n"
            "[BR-08] Confidence profile derived from 5-dimension evaluation scoring "
            "(technical_accuracy, completeness, actionability, consistency, style_match); "
            "drafts scoring ≤3.0 average are flagged for reviewer attention\n"
            "[ASSUMPTION] BR-08 score threshold for reviewer attention — exact value TBC"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 9,
        "epic": "Evaluate each draft",
        "feature": "Read draft alongside comment",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-009",
        "story": (
            "When I open a comment for review, I want to see the original SO comment and "
            "the AI-generated draft response together, "
            "so I can immediately evaluate whether the draft is responsive to the comment."
        ),
        "ac": (
            "[STORY] Original comment and draft response are visible at the same time; "
            "no switching required\n"
            "[BR-01] Intent type (action_request, design_clarification, acknowledgment, "
            "factual_query) is surfaced alongside the draft; "
            "reviewer can orient scrutiny to comment type"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 10,
        "epic": "Evaluate each draft",
        "feature": "Confidence score and failure modes",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-010",
        "story": (
            "When I am reviewing a draft, I want to see the AI's confidence score "
            "(across all five evaluation dimensions) alongside the draft, "
            "so I can calibrate how deeply I need to verify the content."
        ),
        "ac": (
            "[STORY] Overall score and per-dimension breakdown are visible alongside the draft; "
            "failure modes are named, not just counted\n"
            "[BR-08] Per-dimension score breakdown (technical_accuracy, completeness, "
            "actionability, consistency, style_match) visible alongside overall score; "
            "failure mode names shown with plain-language descriptions\n"
            "[ASSUMPTION] BR-08 score threshold calibration values — TBC"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 11,
        "epic": "Evaluate each draft",
        "feature": "Confidence score and failure modes",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-011",
        "story": (
            "When I see a failure mode flag on a draft, I want to understand what the "
            "failure mode means (e.g. insufficient_context vs. wrong_retrieval), "
            "so I know what kind of problem to look for when I read the draft."
        ),
        "ac": (
            "[STORY] Failure mode names are accompanied by brief plain-language descriptions; "
            "the likely consequence is communicated"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 12,
        "epic": "Evaluate each draft",
        "feature": "Inspect source evidence",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-012",
        "story": (
            "When I am reviewing a draft, I want to see which source documents were retrieved "
            "and how relevant each was, "
            "so I can verify that the AI drew on the right evidence."
        ),
        "ac": (
            "[STORY] Retrieved documents are listed with relevance indicators; "
            "the specific passages cited are accessible without leaving the review context\n"
            "[BR-04] Retrieved documents include retrieval method indicator "
            "(semantic / tag-filtered / subsystem-boosted) per source; "
            "relevance score shown per document"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 13,
        "epic": "Evaluate each draft",
        "feature": "Inspect source evidence",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-013",
        "story": (
            "When I am reviewing a draft that references equipment tags, I want to see how "
            "the KKS code was decoded into a subsystem name, "
            "so I can confirm the retrieval was anchored to the correct equipment context."
        ),
        "ac": (
            "[STORY] KKS decoding is visible (code → subsystem name) in the source evidence; "
            "tag-filtered retrieval is distinguishable from semantic retrieval\n"
            "[BR-02] KKS decode path (code → subsystem name) is visible in source evidence; "
            "tag-filtered retrieval is distinguishable from semantic retrieval\n"
            "[ASSUMPTION] BR-02 KKS version reconciliation at ingest time — TBC"
        ),
        "priority": "Low",
        "release": "Slice 3 — Advanced Trust Evaluation",
    },
    {
        "sno": 14,
        "epic": "Evaluate each draft",
        "feature": "Inspect source evidence",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-014",
        "story": (
            "When I am reviewing a draft for a comment from a later revision (R01, R02), "
            "I want to see whether the system retrieved and injected prior revision replies as context, "
            "so I know whether consistency with prior positions was considered."
        ),
        "ac": (
            "[STORY] Revision history injection is visible as a source type; "
            "prior reply text is accessible for comparison\n"
            "[BR-07] Revision history injection is visible as a source type when prior revision "
            "replies were semantically related; prior reply text accessible for comparison"
        ),
        "priority": "Low",
        "release": "Slice 3 — Advanced Trust Evaluation",
    },
    {
        "sno": 15,
        "epic": "Evaluate each draft",
        "feature": "Max-iterations alert",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-015",
        "story": (
            "When a draft was accepted after max grounding iterations, I want to be clearly "
            "warned that some citations may not be verified, "
            "so I know to manually check those references before accepting the draft."
        ),
        "ac": (
            "[STORY] Max-iterations state is prominently indicated; "
            "the engineer knows they bear verification responsibility for those citations\n"
            "[BR-05] When max grounding iterations reached, warning is prominently surfaced; "
            "reviewer is informed they bear citation verification responsibility\n"
            "[ASSUMPTION] BR-05 FAIL_GROUNDING exact detection logic — TBC"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 16,
        "epic": "Evaluate each draft",
        "feature": "Compliance disposition check",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-016",
        "story": (
            "When I am reviewing a draft on a comment that requests a design change, "
            "I want to see whether the AI's draft is defending the existing design or has "
            "incorrectly agreed to the change, so I can detect compliance bias before accepting."
        ),
        "ac": (
            "[STORY] The AI's disposition is surfaced explicitly (defends design / agrees to change); "
            "a compliance-bias flag is present when wrong_disposition is detected\n"
            "[BR-06] AI disposition is surfaced explicitly (defends_design / compliance_bias_flagged); "
            "wrong_disposition triggers compliance-bias flag visible before accept\n"
            "[ASSUMPTION] BR-06 compliance bias detection heuristics — exact detection patterns TBC"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 17,
        "epic": "Evaluate each draft",
        "feature": "Compliance disposition check",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-017",
        "story": (
            "When a compliance-bias flag is present, I want to confirm or override the AI's "
            "disposition before I can complete my review of that draft, "
            "so the flag cannot be silently ignored."
        ),
        "ac": (
            "[STORY] The compliance disposition check is a mandatory step when the flag is present; "
            "accepting without confirming disposition is not possible\n"
            "[BR-06] Compliance disposition check is a mandatory gate when flag is present; "
            "system prevents accepting without confirming or overriding disposition"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 18,
        "epic": "Decide on each draft",
        "feature": "Accept without changes",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-018",
        "story": (
            "When I have reviewed a draft and am satisfied with its content and disposition, "
            "I want to accept it without modification, "
            "so I can move to the next comment efficiently."
        ),
        "ac": (
            "[STORY] Acceptance requires one deliberate action; "
            "the draft is marked as ready and I return to the batch overview"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 19,
        "epic": "Decide on each draft",
        "feature": "Edit draft directly",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-019",
        "story": (
            "When a draft needs minor correction, I want to edit the response text directly, "
            "so I can improve it without needing to regenerate."
        ),
        "ac": (
            "[STORY] I can modify the draft text; "
            "the edited version is saved and marked as ready when I confirm"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 20,
        "epic": "Decide on each draft",
        "feature": "Co-author with AI",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-020",
        "story": (
            "When a draft requires more than minor edits and I need substantive help "
            "generating a better response, I want to use a conversational interface to provide "
            "direction and receive an AI-revised draft, so I am not authoring from scratch."
        ),
        "ac": (
            "[STORY] I can direct the AI conversationally (e.g. 'focus on the safety case "
            "for the current design'); AI produces a revised draft I can review and accept "
            "or continue to iterate"
        ),
        "priority": "Low",
        "release": "Slice 3 — Advanced Trust Evaluation",
    },
    {
        "sno": 21,
        "epic": "Decide on each draft",
        "feature": "Override compliance disposition",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-021",
        "story": (
            "When I have confirmed that the AI draft has incorrectly agreed to an SO change request, "
            "I want to override the disposition and edit the response to defend the existing design, "
            "so the submitted response correctly represents the contractor's position."
        ),
        "ac": (
            "[STORY] Override action is explicit and recorded; "
            "the compliance-bias flag is cleared only when override is confirmed; "
            "edited response reflects design-defense position\n"
            "[BR-06] Override action is explicitly recorded; "
            "compliance-bias flag is cleared only after override confirmed; "
            "edited response must reflect design-defense position"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 22,
        "epic": "Decide on each draft",
        "feature": "Escalate to domain expert",
        "screen": "OV-03 Escalation Overlay",
        "id": "DS-022",
        "story": (
            "When I cannot resolve a comment on my own (domain knowledge required beyond "
            "my expertise), I want to escalate it to a domain expert for review, "
            "so the comment is not submitted with a response I cannot validate."
        ),
        "ac": (
            "[STORY] Escalation marks the comment as held; I can continue reviewing other comments; "
            "the batch cannot be fully submitted until the escalated comment is resolved "
            "or explicitly released"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 23,
        "epic": "Decide on each draft",
        "feature": "Escalate to domain expert",
        "screen": "OV-03 Escalation Overlay",
        "id": "DS-023",
        "story": (
            "When I escalate a comment to a domain expert, I want to include the AI draft "
            "and my concerns as context, so the domain expert has everything they need to "
            "resolve it without starting from scratch."
        ),
        "ac": (
            "[STORY] Escalation includes the AI draft, trust signals, and a notes field "
            "for the reviewer's concerns"
        ),
        "priority": "Medium",
        "release": "Slice 2 — Trust Signal Panel",
    },
    {
        "sno": 24,
        "epic": "Submit the batch",
        "feature": "Confirm all comments reviewed",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-024",
        "story": (
            "When I have finished reviewing all comments in the batch, I want to see a clear "
            "confirmation that every comment has been acted on (accepted, edited, escalated), "
            "so I know the batch is complete before submitting."
        ),
        "ac": (
            "[STORY] Batch cannot be submitted while any comment has an unresolved review state; "
            "unresolved comments are indicated"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 25,
        "epic": "Submit the batch",
        "feature": "Submit for peer review",
        "screen": "OV-01 View Comments Drawer",
        "id": "DS-025",
        "story": (
            "When all comments are reviewed and ready, I want to submit the batch for peer "
            "review as a single action, so the review workflow advances without additional "
            "manual handoffs."
        ),
        "ac": (
            "[STORY] Submission triggers the peer review step; I receive confirmation that "
            "the batch is in peer review; I can see the batch status"
        ),
        "priority": "High",
        "release": "Slice 1 — Basic Review Loop",
    },
    {
        "sno": 26,
        "epic": "Submit the batch",
        "feature": "Track through approval",
        "screen": "P-02 Project Space",
        "id": "DS-026",
        "story": (
            "When a batch is in peer review or domain expert approval, I want to see its "
            "current status, so I know where it is in the process and whether any follow-up "
            "is needed."
        ),
        "ac": (
            "[STORY] Batch status is visible (in peer review / awaiting domain expert approval "
            "/ submitted to client) after submission"
        ),
        "priority": "Low",
        "release": "Slice 3 — Advanced Trust Evaluation",
    },
]

RBAC_ROLES = [
    "Process Engineer",
    "Domain Expert",
    "PMO / Claims Manager",
    "New Engineer",
    "Document Controller",
]

RBAC_ROWS = [
    # (group, screen, PE, DE, PMO, NE, DC)
    ("Product Navigation", "P-01 Projects Home",              "Full",             "Full",                   "Full",       "Full",       "Full"),
    ("Product Navigation", "P-02 Project Space",              "Full",             "Read-only",              "Read-only",  "Read-only",  "Full"),
    ("Product Navigation", "P-03 Pending Reviews",            "Full (Review tab)","Full (Approve tab)",     "No access",  "No access",  "No access"),
    ("Product Navigation", "P-04 Ask",                        "Full",             "Full",                   "No access",  "Full",       "No access"),
    ("Product Navigation", "P-05 Knowledge Studio",           "Read-only",        "Full",                   "No access",  "Read-only",  "No access"),
    ("Overlay",            "OV-01 View Comments Drawer",      "Full",             "Read-only",              "No access",  "Read-only",  "No access"),
    ("Overlay",            "OV-02 Knowledge Gap Review",      "No access",        "Full",                   "No access",  "No access",  "No access"),
    ("Overlay",            "OV-03 Escalation Overlay",        "Full (initiate)",  "No access",              "No access",  "No access",  "No access"),
    ("Domain Expert",      "DE-01 Escalation Queue",          "No access",        "Full",                   "No access",  "No access",  "No access"),
    ("Domain Expert",      "DE-02 Escalation Detail",         "No access",        "Full",                   "No access",  "No access",  "No access"),
]

DATA_FIELDS = [
    # (entity, field_name, label, data_type, mandatory, editable, visible_to_roles, validation, lov_ref, remarks)
    ("RFA",           "rfa_id",          "RFA ID",               "String",   "Yes", "No",  "All roles",           "",                           "",                    "Primary key"),
    ("RFA",           "title",           "RFA Title",            "String",   "Yes", "No",  "All roles",           "",                           "",                    ""),
    ("RFA",           "revision_number", "Revision No.",         "Integer",  "Yes", "No",  "All roles",           "Positive integer",           "",                    ""),
    ("RFA",           "submission_date", "Submission Date",      "Date",     "Yes", "No",  "All roles",           "Valid date",                 "",                    ""),
    ("RFA",           "status",          "Status",               "Enum",     "Yes", "Yes", "All roles",           "",                           "Batch Status",        ""),
    ("Revision",      "revision_id",     "Revision ID",          "String",   "Yes", "No",  "All roles",           "",                           "",                    "Primary key"),
    ("Revision",      "rfa_id",          "RFA ID",               "String",   "Yes", "No",  "All roles",           "FK → RFA.rfa_id",            "",                    "Foreign key"),
    ("Revision",      "submission_date", "Submission Date",      "Date",     "Yes", "No",  "All roles",           "Valid date",                 "",                    ""),
    ("Revision",      "prior_revision_ids","Prior Revision IDs", "Array",    "No",  "No",  "All roles",           "",                           "",                    "Empty for R00"),
    ("Comment",       "comment_id",      "Comment ID",           "String",   "Yes", "No",  "All roles",           "",                           "",                    "Primary key"),
    ("Comment",       "rfa_id",          "RFA ID",               "String",   "Yes", "No",  "All roles",           "FK → RFA.rfa_id",            "",                    ""),
    ("Comment",       "revision_id",     "Revision ID",          "String",   "Yes", "No",  "All roles",           "FK → Revision.revision_id",  "",                    ""),
    ("Comment",       "author_org",      "Submitting Organisation","String", "Yes", "No",  "All roles",           "",                           "",                    ""),
    ("Comment",       "intent_type",     "Intent Type",          "Enum",     "Yes", "No",  "All roles",           "",                           "Intent Type",         "Classified by BR-01"),
    ("Comment",       "raw_text",        "Comment Text",         "Text",     "Yes", "No",  "All roles",           "",                           "",                    ""),
    ("Comment",       "review_state",    "Review State",         "Enum",     "Yes", "Yes", "Process Engineer, Domain Expert","",               "Comment Review State", ""),
    ("Draft Response","draft_id",        "Draft ID",             "String",   "Yes", "No",  "All roles",           "",                           "",                    "Primary key"),
    ("Draft Response","comment_id",      "Comment ID",           "String",   "Yes", "No",  "All roles",           "FK → Comment.comment_id",    "",                    ""),
    ("Draft Response","response_text",   "Draft Response Text",  "Text",     "Yes", "Yes", "Process Engineer",    "",                           "",                    "Editable by reviewer"),
    ("Draft Response","disposition",     "Compliance Disposition","Enum",    "Yes", "Yes", "Process Engineer",    "",                           "Compliance Disposition",""),
    ("Draft Response","validation_state","Validation State",     "Enum",     "Yes", "No",  "Process Engineer",    "",                           "Validation State",    "Set by AI pipeline"),
    ("Draft Response","generation_method","Generation Method",   "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    ""),
    ("Draft Response","iteration_count", "Iteration Count",      "Integer",  "Yes", "No",  "Process Engineer",    "≥1",                         "",                    "Max iterations per BR-05"),
    ("Trust Signal",  "trust_signal_id", "Trust Signal ID",      "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    "Primary key"),
    ("Trust Signal",  "draft_id",        "Draft ID",             "String",   "Yes", "No",  "Process Engineer",    "FK → Draft Response.draft_id","",                  ""),
    ("Trust Signal",  "overall_score",   "Overall Score",        "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "Average of 5 dimensions"),
    ("Trust Signal",  "technical_accuracy","Technical Accuracy", "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "BR-08 dimension"),
    ("Trust Signal",  "completeness",    "Completeness",         "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "BR-08 dimension"),
    ("Trust Signal",  "actionability",   "Actionability",        "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "BR-08 dimension"),
    ("Trust Signal",  "consistency",     "Consistency",          "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "BR-08 dimension"),
    ("Trust Signal",  "style_match",     "Style Match",          "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–5.0",                    "",                    "BR-08 dimension"),
    ("Trust Signal",  "failure_modes",   "Failure Modes",        "Array",    "No",  "No",  "Process Engineer",    "",                           "Failure Mode",        "Empty if no failures"),
    ("Source Evidence","evidence_id",    "Evidence ID",          "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    "Primary key"),
    ("Source Evidence","draft_id",       "Draft ID",             "String",   "Yes", "No",  "Process Engineer",    "FK → Draft Response.draft_id","",                  ""),
    ("Source Evidence","document_id",    "Document ID",          "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    "Links to DMS"),
    ("Source Evidence","document_title", "Document Title",       "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    ""),
    ("Source Evidence","relevance_score","Relevance Score",      "Decimal",  "Yes", "No",  "Process Engineer",    "0.0–1.0",                    "",                    ""),
    ("Source Evidence","retrieval_method","Retrieval Method",    "Enum",     "Yes", "No",  "Process Engineer",    "",                           "Retrieval Method",    "BR-04"),
    ("Source Evidence","cited_passages", "Cited Passages",       "Array",    "No",  "No",  "Process Engineer",    "",                           "",                    ""),
    ("KKS Code",      "kks_code",        "KKS Code",             "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    "Primary key; BR-02"),
    ("KKS Code",      "equipment_name",  "Equipment Name",       "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    ""),
    ("KKS Code",      "subsystem",       "Subsystem",            "String",   "Yes", "No",  "Process Engineer",    "",                           "",                    ""),
    ("KKS Code",      "parent_system",   "Parent System",        "String",   "No",  "No",  "Process Engineer",    "",                           "",                    ""),
    ("Batch",         "batch_id",        "Batch ID",             "String",   "Yes", "No",  "All roles",           "",                           "",                    "Primary key"),
    ("Batch",         "rfa_id",          "RFA ID",               "String",   "Yes", "No",  "All roles",           "FK → RFA.rfa_id",            "",                    ""),
    ("Batch",         "revision_id",     "Revision ID",          "String",   "Yes", "No",  "All roles",           "FK → Revision.revision_id",  "",                    ""),
    ("Batch",         "status",          "Batch Status",         "Enum",     "Yes", "Yes", "All roles",           "",                           "Batch Status",        ""),
    ("Batch",         "total_comments",  "Total Comments",       "Integer",  "Yes", "No",  "All roles",           "≥0",                         "",                    ""),
    ("Batch",         "flagged_count",   "Flagged Count",        "Integer",  "Yes", "No",  "All roles",           "≥0",                         "",                    ""),
    ("Batch",         "submitted_at",    "Submitted At",         "Datetime", "No",  "No",  "All roles",           "",                           "",                    "Null until submitted"),
    ("Escalation",    "escalation_id",   "Escalation ID",        "String",   "Yes", "No",  "Process Engineer, Domain Expert","",               "",                    "Primary key"),
    ("Escalation",    "comment_id",      "Comment ID",           "String",   "Yes", "No",  "Process Engineer, Domain Expert","FK → Comment.comment_id","",            ""),
    ("Escalation",    "draft_id",        "Draft ID",             "String",   "Yes", "No",  "Process Engineer, Domain Expert","FK → Draft Response.draft_id","",        ""),
    ("Escalation",    "escalated_by",    "Escalated By",         "String",   "Yes", "No",  "Process Engineer, Domain Expert","",               "",                    "User ID"),
    ("Escalation",    "assigned_to",     "Assigned To",          "String",   "Yes", "Yes", "Domain Expert",       "",                           "",                    "Domain Expert user ID"),
    ("Escalation",    "notes",           "Escalation Notes",     "Text",     "No",  "Yes", "Process Engineer, Domain Expert","",               "",                    "Reviewer's concerns"),
    ("Escalation",    "status",          "Escalation Status",    "String",   "Yes", "Yes", "Process Engineer, Domain Expert","",               "",                    ""),
    ("Escalation",    "resolved_at",     "Resolved At",          "Datetime", "No",  "No",  "Process Engineer, Domain Expert","",               "",                    "Null until resolved"),
]

LOV_ROWS = [
    # (list_name, display_value, code, description, active)
    ("Intent Type",         "Action Request",              "action_request",         "Comment requests a change to the design",                              "Yes"),
    ("Intent Type",         "Design Clarification",        "design_clarification",   "Comment requests clarification without demanding a change",           "Yes"),
    ("Intent Type",         "Acknowledgment",              "acknowledgment",         "Comment acknowledges receipt or confirms understanding",              "Yes"),
    ("Intent Type",         "Factual Query",               "factual_query",          "Comment asks a factual question about the system or design",          "Yes"),
    ("Failure Mode",        "Incomplete Response",         "incomplete_response",    "Draft response does not fully address all aspects of the comment",    "Yes"),
    ("Failure Mode",        "Misunderstood Intent",        "misunderstood_intent",   "AI misclassified the comment intent type",                            "Yes"),
    ("Failure Mode",        "Insufficient Context",        "insufficient_context",   "Retrieved documents did not provide enough grounding for the response","Yes"),
    ("Failure Mode",        "Wrong Disposition",           "wrong_disposition",      "AI agreed to a change request instead of defending the design (BR-06)","Yes"),
    ("Failure Mode",        "Wrong Retrieval",             "wrong_retrieval",        "Retrieved documents were not relevant to the comment",                "Yes"),
    ("Failure Mode",        "Undercommunication",          "undercommunication",     "Response omits key technical detail or qualification",                "Yes"),
    ("Retrieval Method",    "Semantic Search",             "semantic",               "Vector similarity search across the knowledge corpus",               "Yes"),
    ("Retrieval Method",    "Tag-Filtered",                "tag_filtered",           "Search filtered by KKS equipment tag (BR-02)",                       "Yes"),
    ("Retrieval Method",    "Subsystem-Boosted",           "subsystem_boosted",      "Semantic search with relevance boost for the matched subsystem",      "Yes"),
    ("Validation State",    "Grounding Passed",            "grounding_passed",       "All citations validated within allowed iterations",                   "Yes"),
    ("Validation State",    "Grounding Failed — Retry",    "fail_grounding_retry",   "Grounding check failed; retry attempted",                             "Yes"),
    ("Validation State",    "Max Iterations — Accepted",   "max_iterations_accepted","Grounding failed after max retries; draft accepted with warning (BR-05)","Yes"),
    ("Compliance Disposition","Defends Design",            "defends_design",         "Draft correctly defends the existing design against the change request","Yes"),
    ("Compliance Disposition","Compliance Bias Flagged",   "compliance_bias_flagged","AI incorrectly agreed to a change request — reviewer action required (BR-06)","Yes"),
    ("Compliance Disposition","Override Confirmed",        "override_confirmed",     "Reviewer confirmed and overrode the AI's wrong disposition",          "Yes"),
    ("Comment Review State","Not Yet Reviewed",            "unreviewed",             "Comment has not been opened by the reviewer",                         "Yes"),
    ("Comment Review State","Accepted",                    "accepted",               "Reviewer accepted the AI draft without changes",                      "Yes"),
    ("Comment Review State","Edited",                      "edited",                 "Reviewer edited the AI draft before accepting",                       "Yes"),
    ("Comment Review State","Override Applied",            "override_applied",       "Reviewer overrode compliance disposition and edited the response",    "Yes"),
    ("Comment Review State","Escalated",                   "escalated",              "Comment escalated to Domain Expert; batch held pending resolution",   "Yes"),
    ("Comment Review State","Held for Expert",             "held_for_expert",        "Comment held for domain expert review; not yet assigned",             "Yes"),
    ("Batch Status",        "Draft Generated",             "draft_generated",        "AI pipeline has completed; drafts ready for review",                  "Yes"),
    ("Batch Status",        "In Review",                   "in_review",              "Process Engineer actively reviewing comments",                        "Yes"),
    ("Batch Status",        "In Peer Review",              "in_peer_review",         "Batch submitted by reviewer; awaiting peer review",                   "Yes"),
    ("Batch Status",        "Awaiting Expert Approval",    "awaiting_expert_approval","Batch passed peer review; awaiting Domain Expert sign-off",          "Yes"),
    ("Batch Status",        "Submitted to Client",         "submitted_to_client",    "Batch fully approved and delivered to the submitting organisation",   "Yes"),
    ("Priority",            "High",                        "High",                   "Walking skeleton stories; must be in Slice 1",                        "Yes"),
    ("Priority",            "Medium",                      "Medium",                 "Trust signal and evidence stories; Slice 2",                          "Yes"),
    ("Priority",            "Low",                         "Low",                    "Advanced evaluation and co-author features; Slice 3",                 "Yes"),
]

NOTIFICATION_ROWS = [
    # (sno, story_ids, trigger, triggered_by, recipients, in_app, email, sms, push, remarks)
    (1, "DS-001, DS-002, DS-003",
     "AI pipeline completes processing a revision batch — drafts are ready for review",
     "AI pipeline / DMS ingest",
     "Process Engineer",
     "✓", "✗", "✗", "✗",
     "Routes to P-02 Project Space with relevant RFA pre-highlighted; "
     "notification content: 'New batch ready: [RFA ID] — [N] comments, [N] flagged'"),
    (2, "DS-022, DS-023",
     "Process Engineer escalates a comment to Domain Expert",
     "OV-03 Escalation Overlay",
     "Domain Expert",
     "✓", "✗", "✗", "✗",
     "Routes to DE-01 Escalation Queue; "
     "content: 'Comment escalated: [Document name] §[section]'"),
    (3, "DS-022, DS-023",
     "Domain Expert resolves an escalated comment",
     "DE-02 Escalation Detail",
     "Process Engineer",
     "✓", "✗", "✗", "✗",
     "Routes back to OV-01 View Comments Drawer for the resolved comment; "
     "content: 'Escalation resolved: [Document name] §[section]'"),
    (4, "DS-025",
     "Batch submitted for peer review",
     "OV-01 View Comments Drawer (submit action)",
     "Peer Reviewer",
     "✓", "✗", "✗", "✗",
     "Routes to P-03 Pending Reviews (Approve tab); "
     "content: 'Batch submitted for peer review: [RFA ID], [N] comments'"),
    (5, "DS-015",
     "Max grounding iterations reached on a draft response",
     "AI pipeline",
     "Process Engineer (inline)",
     "✗", "✗", "✗", "✗",
     "Inline warning badge within OV-01 View Comments Drawer; "
     "no separate push notification — reviewer sees indicator when opening the draft"),
    (6, "DS-007, DS-016, DS-017",
     "Compliance bias (wrong_disposition) detected on a draft response",
     "AI pipeline",
     "Process Engineer (inline)",
     "✗", "✗", "✗", "✗",
     "Inline compliance flag badge within OV-01 View Comments Drawer; "
     "no separate push notification — reviewer sees indicator in batch overview and on draft"),
]

# ---------------------------------------------------------------------------
# WRITE
# ---------------------------------------------------------------------------

def populate():
    wb = load_workbook(BRD_PATH)

    # --- User Stories ---
    ws = wb["User Stories"]
    for s in STORIES:
        r = s["sno"] + 3  # row 4 = sno 1
        ws.cell(r, 1).value = s["sno"]
        ws.cell(r, 2).value = s["epic"]
        ws.cell(r, 3).value = s["screen"]
        ws.cell(r, 4).value = s["id"]
        ws.cell(r, 5).value = s["story"]
        ws.cell(r, 6).value = s["ac"]
        ws.cell(r, 7).value = s["priority"]
        ws.cell(r, 8).value = s["release"]
    print(f"  User Stories: wrote {len(STORIES)} rows (rows 4–{3+len(STORIES)})")

    # --- RBAC — update role name headers then write data ---
    ws2 = wb["RBAC"]
    for i, role in enumerate(RBAC_ROLES):
        ws2.cell(3, 4 + i).value = role
    for i, row in enumerate(RBAC_ROWS, start=1):
        r = i + 3  # data starts row 4
        ws2.cell(r, 1).value = i
        ws2.cell(r, 2).value = row[0]
        ws2.cell(r, 3).value = row[1]
        ws2.cell(r, 4).value = row[2]
        ws2.cell(r, 5).value = row[3]
        ws2.cell(r, 6).value = row[4]
        ws2.cell(r, 7).value = row[5]
        ws2.cell(r, 8).value = row[6]
    print(f"  RBAC: wrote {len(RBAC_ROWS)} rows")

    # --- Data Fields ---
    ws3 = wb["Data Fields"]
    for i, row in enumerate(DATA_FIELDS, start=1):
        r = i + 2  # header in row 2, data from row 3
        ws3.cell(r, 1).value = i
        ws3.cell(r, 2).value = row[0]   # Module / Screen (entity name)
        ws3.cell(r, 3).value = row[1]   # Field Name
        ws3.cell(r, 4).value = row[2]   # Field Label (Display)
        ws3.cell(r, 5).value = row[3]   # Data Type
        ws3.cell(r, 6).value = row[4]   # Mandatory
        ws3.cell(r, 7).value = row[5]   # Editable
        ws3.cell(r, 8).value = row[6]   # Visible To Roles
        ws3.cell(r, 9).value = row[7]   # Validation Rules
        ws3.cell(r, 10).value = row[8]  # LOV Reference
        ws3.cell(r, 11).value = row[9]  # Remarks
    print(f"  Data Fields: wrote {len(DATA_FIELDS)} rows")

    # --- LOV ---
    ws4 = wb["LOV"]
    for i, row in enumerate(LOV_ROWS, start=1):
        r = i + 3  # headers in row 3, data from row 4
        ws4.cell(r, 1).value = i
        ws4.cell(r, 2).value = row[0]   # List Name
        ws4.cell(r, 3).value = row[1]   # Value (Display)
        ws4.cell(r, 4).value = row[2]   # Code (System)
        ws4.cell(r, 5).value = row[3]   # Description / Meaning
        ws4.cell(r, 6).value = row[4]   # Active
    print(f"  LOV: wrote {len(LOV_ROWS)} rows")

    # --- Notification Mapping ---
    ws5 = wb["Notification Mapping"]
    for row in NOTIFICATION_ROWS:
        r = row[0] + 2  # header in row 2, data from row 3
        ws5.cell(r, 1).value = row[0]   # S/No.
        ws5.cell(r, 2).value = row[1]   # User Story ID
        ws5.cell(r, 3).value = row[2]   # Notification Trigger
        ws5.cell(r, 4).value = row[3]   # Triggered By
        ws5.cell(r, 5).value = row[4]   # Recipient(s)
        ws5.cell(r, 6).value = row[5]   # In-App
        ws5.cell(r, 7).value = row[6]   # Email
        ws5.cell(r, 8).value = row[7]   # SMS
        ws5.cell(r, 9).value = row[8]   # Push
        ws5.cell(r, 10).value = row[9]  # Remarks
    print(f"  Notification Mapping: wrote {len(NOTIFICATION_ROWS)} rows")

    wb.save(BRD_PATH)
    print(f"\nSaved: {BRD_PATH}")


def recalc():
    if not os.path.exists(RECALC):
        print(f"  recalc.py not found at {RECALC} — skipping formula recalculation")
        return
    result = subprocess.run(
        [sys.executable, RECALC, BRD_PATH],
        capture_output=True, text=True
    )
    print(f"  recalc output: {result.stdout.strip()}")
    if result.returncode != 0:
        print(f"  recalc stderr: {result.stderr.strip()}")


if __name__ == "__main__":
    print("=== Populating design/BRD.xlsx ===\n")
    populate()
    print("\nRecalculating formulas...")
    recalc()
    print("\nDone.")
